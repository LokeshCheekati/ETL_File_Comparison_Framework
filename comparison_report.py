"""Generates one detailed Excel report for one comparison."""

from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from logger import logger


def _clean_source_only(data_df):
    if data_df.empty:
        return data_df
    data_df = data_df.drop(
        columns=[c for c in data_df.columns if c.endswith("_TARGET") or c == "_merge"],
        errors="ignore"
    )
    data_df.columns = [c.replace("_SOURCE", "") for c in data_df.columns]
    return data_df


def _clean_target_only(data_df):
    if data_df.empty:
        return data_df
    data_df = data_df.drop(
        columns=[c for c in data_df.columns if c.endswith("_SOURCE") or c == "_merge"],
        errors="ignore"
    )
    data_df.columns = [c.replace("_TARGET", "") for c in data_df.columns]
    return data_df


def _format_workbook(writer):
    bold_font = Font(bold=True)
    green_fill = PatternFill(fill_type="solid", start_color="C6EFCE")
    red_fill = PatternFill(fill_type="solid", start_color="FFC7CE")

    for sheet in writer.sheets.values():
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = bold_font
        for column_cells in sheet.columns:
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=0
            )
            column_letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    summary_sheet = writer.sheets.get(config.SUMMARY_SHEET)
    if summary_sheet:
        status_column = next(
            (cell.column for cell in summary_sheet[1] if cell.value == "STATUS"),
            None
        )
        if status_column:
            for row in range(2, summary_sheet.max_row + 1):
                cell = summary_sheet.cell(row=row, column=status_column)
                if cell.value == "PASS":
                    cell.fill = green_fill
                elif cell.value == "FAIL":
                    cell.fill = red_fill


def generate_comparison_report(result):
    """Generate a separate workbook for one comparison."""
    comparison_name = result["comparison_name"]
    logger.info(f"Generating comparison report: {comparison_name}")

    summary_df = pd.DataFrame([{
        "COMPARISON_NAME": comparison_name,
        "SOURCE_FILE": result["source_file"],
        "TARGET_FILE": result["target_file"],
        "SOURCE_COUNT": result["source_count"],
        "TARGET_COUNT": result["target_count"],
        "SOURCE_ONLY_COUNT": len(result["source_only_df"]),
        "TARGET_ONLY_COUNT": len(result["target_only_df"]),
        "MISMATCH_COUNT": len(result["mismatch_df"]),
        "STATUS": result["status"],
    }])

    source_only_df = _clean_source_only(result["source_only_df"].copy())
    target_only_df = _clean_target_only(result["target_only_df"].copy())
    mismatch_df = result["mismatch_df"].copy()

    safe_name = "".join(
        c if c.isalnum() or c in "._-" else "_"
        for c in comparison_name
    )

    Path(config.OUTPUT_FOLDER).mkdir(parents=True, exist_ok=True)

    report_name = (
        f"{config.REPORT_NAME}_{safe_name}_"
        f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    )
    report_path = Path(config.OUTPUT_FOLDER) / report_name

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=config.SUMMARY_SHEET, index=False)
        source_only_df.to_excel(writer, sheet_name=config.SOURCE_ONLY_SHEET, index=False)
        target_only_df.to_excel(writer, sheet_name=config.TARGET_ONLY_SHEET, index=False)

        if len(mismatch_df) > 1_048_576:
            csv_name = (
                f"{config.REPORT_NAME}_{safe_name}_"
                f"{datetime.now():%Y%m%d_%H%M%S}_Data_Mismatches.csv"
            )
            csv_path = Path(config.OUTPUT_FOLDER) / csv_name
            mismatch_df.to_csv(csv_path, index=False)
            logger.warning(
                f"Mismatch sheet exceeds Excel row limit. "
                f"Writing full mismatch data as CSV: {csv_path}"
            )
            pd.DataFrame([{
                "MESSAGE": f"Full mismatch data is available in {csv_path.name}"
            }]).to_excel(writer, sheet_name=config.MISMATCH_SHEET, index=False)
        else:
            mismatch_df.to_excel(writer, sheet_name=config.MISMATCH_SHEET, index=False)

        _format_workbook(writer)

    logger.info(f"Comparison report generated successfully: {report_path}")
    return report_path
