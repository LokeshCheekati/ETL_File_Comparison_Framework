"""Generates one compact summary Excel report for the complete run."""

from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from logger import logger


def _format_summary(writer):
    bold_font = Font(bold=True)
    green_fill = PatternFill(fill_type="solid", start_color="C6EFCE")
    red_fill = PatternFill(fill_type="solid", start_color="FFC7CE")

    for sheet in writer.sheets.values():
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = bold_font
        for column_cells in sheet.columns:
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=0
            )
            letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[letter].width = min(max_length + 2, 60)

    sheet = writer.sheets[config.SUMMARY_SHEET]
    status_column = next(
        (cell.column for cell in sheet[1] if cell.value == "STATUS"),
        None
    )
    if status_column:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=status_column)
            if cell.value == "PASS":
                cell.fill = green_fill
            elif cell.value == "FAIL":
                cell.fill = red_fill


def generate_summary_report(all_results):
    logger.info("Generating summary report.")

    summary_records = []
    for result in all_results:
        summary_records.append({
            "COMPARISON_NAME": result["comparison_name"],
            "SOURCE_FILE": result["source_file"],
            "TARGET_FILE": result["target_file"],
            "SOURCE_COUNT": result["source_count"],
            "TARGET_COUNT": result["target_count"],
            "SOURCE_ONLY_COUNT": len(result["source_only_df"]),
            "TARGET_ONLY_COUNT": len(result["target_only_df"]),
            "MISMATCH_COUNT": len(result["mismatch_df"]),
            "STATUS": result["status"],
        })

    summary_df = pd.DataFrame(summary_records)
    Path(config.OUTPUT_FOLDER).mkdir(parents=True, exist_ok=True)

    report_name = (
        f"{config.SUMMARY_REPORT_NAME}_"
        f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    )
    report_path = Path(config.OUTPUT_FOLDER) / report_name

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=config.SUMMARY_SHEET, index=False)
        _format_summary(writer)

    logger.info(f"Summary report generated successfully: {report_path}")
    return report_path
