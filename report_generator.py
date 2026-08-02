"""
Generates the Excel comparison report containing summary,
missing records and data mismatches.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import config
from logger import logger
from datetime import datetime


def generate_report(all_results):
    """
    Generates the comparison report in Excel format.
    """

    logger.info("Generating comparison report.")

    summary_records = []
    source_only_df_list = []
    target_only_df_list = []
    mismatch_df_list = []

    for result in all_results:

        summary_records.append(
            {
                "COMPARISON_NAME": result["comparison_name"],
                "SOURCE_COUNT": result["source_count"],
                "TARGET_COUNT": result["target_count"],
                "SOURCE_ONLY_COUNT": len(result["source_only_df"]),
                "TARGET_ONLY_COUNT": len(result["target_only_df"]),
                "MISMATCH_COUNT": len(result["mismatch_df"]),
                "STATUS": result["status"]
            }
        )

        # Collect source only records.
        if not result["source_only_df"].empty:
            temp_df = result["source_only_df"].copy()
            temp_df.insert(
                0,
                "COMPARISON_NAME",
                result["comparison_name"]
            )
            source_only_df_list.append(temp_df)

        # Collect target only records.
        if not result["target_only_df"].empty:
            temp_df = result["target_only_df"].copy()
            temp_df.insert(
                0,
                "COMPARISON_NAME",
                result["comparison_name"]
            )
            target_only_df_list.append(temp_df)

        # Collect data mismatches.
        if not result["mismatch_df"].empty:
            mismatch_df_list.append(result["mismatch_df"])

    summary_df = pd.DataFrame(summary_records)

    source_only_df = (
        pd.concat(source_only_df_list, ignore_index=True)
        if source_only_df_list
        else pd.DataFrame()
    )

    target_only_df = (
        pd.concat(target_only_df_list, ignore_index=True)
        if target_only_df_list
        else pd.DataFrame()
    )

    mismatch_df = (
        pd.concat(mismatch_df_list, ignore_index=True)
        if mismatch_df_list
        else pd.DataFrame()
    )

    # Remove unnecessary columns from Missing in Source.
    if not source_only_df.empty:

        source_only_df = source_only_df.drop(
            columns=[
                col
                for col in source_only_df.columns
                if col.endswith("_TARGET") or col == "_merge"
            ]
        )

        source_only_df.columns = [
            col.replace("_SOURCE", "")
            for col in source_only_df.columns
        ]

    # Remove unnecessary columns from Missing in Target.
    if not target_only_df.empty:

        target_only_df = target_only_df.drop(
            columns=[
                col
                for col in target_only_df.columns
                if col.endswith("_SOURCE") or col == "_merge"
            ]
        )

        target_only_df.columns = [
            col.replace("_TARGET", "")
            for col in target_only_df.columns
        ]

    report_name = (
    f"{config.REPORT_NAME}_"
    f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    )

    report_path = Path(config.OUTPUT_FOLDER) / report_name

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:

        summary_df.to_excel(
            writer,
            sheet_name=config.SUMMARY_SHEET,
            index=False
        )

        source_only_df.to_excel(
            writer,
            sheet_name=config.SOURCE_ONLY_SHEET,
            index=False
        )

        target_only_df.to_excel(
            writer,
            sheet_name=config.TARGET_ONLY_SHEET,
            index=False
        )

        mismatch_df.to_excel(
            writer,
            sheet_name=config.MISMATCH_SHEET,
            index=False
        )

        # -----------------------------
        # Excel Formatting
        # -----------------------------

        bold_font = Font(bold=True)

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE"
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="FFC7CE"
        )

        for sheet in writer.sheets.values():

            # Freeze header row.
            sheet.freeze_panes = "A2"

            # Make header bold.
            for cell in sheet[1]:
                cell.font = bold_font

            # Auto-adjust column widths.
            for column_cells in sheet.columns:

                max_length = 0
                column_letter = get_column_letter(
                    column_cells[0].column
                )

                for cell in column_cells:

                    try:
                        if cell.value:
                            max_length = max(
                                max_length,
                                len(str(cell.value))
                            )
                    except Exception:
                        pass

                sheet.column_dimensions[
                    column_letter
                ].width = max_length + 2

        # Highlight PASS / FAIL in Summary sheet.
        summary_sheet = writer.sheets[
            config.SUMMARY_SHEET
        ]

        status_column = None

        for cell in summary_sheet[1]:

            if cell.value == "STATUS":
                status_column = cell.column
                break

        if status_column:

            for row in range(
                2,
                summary_sheet.max_row + 1
            ):

                status_cell = summary_sheet.cell(
                    row=row,
                    column=status_column
                )

                if status_cell.value == "PASS":
                    status_cell.fill = green_fill

                elif status_cell.value == "FAIL":
                    status_cell.fill = red_fill

    logger.info(
        f"Report generated successfully: {report_path}"
    )